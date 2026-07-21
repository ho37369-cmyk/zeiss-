"""Write the compact cell-count statistics workbook."""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")


def _excel_value(value, default=""):
    """Convert legacy/numpy/container values into values accepted by Excel."""
    if value is None:
        return default
    if hasattr(value, "item"):
        try:
            value = value.item()
        except (ValueError, TypeError):
            pass
    if isinstance(value, (list, tuple, set, dict)):
        return json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _count_value(value):
    """Return a safe integer count even when a legacy task supplied bad data."""
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def write_excel(filepath, results):
    """Create one sheet containing the requested total/dead cell counts."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "细胞统计"
    headers = ["文件名", "视野", "所有细胞数", "死细胞数"]
    _write_header(ws, headers)

    for ri, r in enumerate(results, start=2):
        vals = [
            _excel_value(r.get("filename", "")),
            _excel_value(r.get("scene", "")),
            _count_value(r.get("total")),
            _count_value(r.get("dead")),
        ]
        for ci, v in enumerate(vals, start=1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = BORDER
            cell.alignment = CENTER

    if len(results) > 1:
        total_row = len(results) + 2
        totals = ["合计", "", sum(_count_value(r.get("total")) for r in results),
                  sum(_count_value(r.get("dead")) for r in results)]
        for ci, value in enumerate(totals, start=1):
            cell = ws.cell(row=total_row, column=ci, value=value)
            cell.font = Font(name="Calibri", bold=True)
            cell.border = BORDER
            cell.alignment = CENTER

    _auto_fit(ws, headers)
    ws.freeze_panes = "A2"
    wb.save(filepath)


def _write_header(ws, headers):
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _auto_fit(ws, headers):
    for ci, h in enumerate(headers, start=1):
        width = sum(2 if ord(c) > 127 else 1 for c in h) + 4
        ws.column_dimensions[get_column_letter(ci)].width = max(width, 12)
