import unittest
import tempfile
from pathlib import Path

import numpy as np
import openpyxl

from cell_counter.annotator import match_dead_cells
from cell_counter.cell_counter import _compact_labels, count_cells
from cell_counter.channel_classifier import classify_channels
from cell_counter.excel_writer import write_excel


class ChannelClassificationTests(unittest.TestCase):
    def test_generic_channels_are_classified_by_image_and_assigned_roles(self):
        rng = np.random.default_rng(7)
        dead = rng.integers(0, 5, (128, 128), dtype=np.uint8)
        dead[30:36, 40:46] = 220
        total = rng.integers(0, 15, (128, 128), dtype=np.uint8)
        for y, x in [(20, 20), (50, 70), (90, 40), (100, 100)]:
            total[y - 5:y + 5, x - 5:x + 5] = 180
        brightfield = np.clip(
            rng.normal(120, 15, (128, 128)), 0, 255).astype(np.uint8)

        scenes = {"Scene0": {"channels": [
            {"index": 0, "name": "Channel 0", "image": dead},
            {"index": 1, "name": "Channel 1", "image": total},
            {"index": 2, "name": "Channel 2", "image": brightfield},
        ]}}
        metadata = [
            {"index": i, "name": f"Channel {i}", "dye": None}
            for i in range(3)
        ]

        result = classify_channels(metadata, scenes)
        self.assertEqual([c.channel_type for c in result],
                         ["fluorescence", "fluorescence", "brightfield"])
        self.assertEqual([c.role for c in result], ["dead", "total", None])

    def test_named_hoechst_is_total_and_pi_is_dead(self):
        image = np.zeros((32, 32), dtype=np.uint8)
        scenes = {"Scene0": {"channels": [
            {"index": 0, "name": "Hoechst", "image": image},
            {"index": 1, "name": "PI", "image": image},
        ]}}
        metadata = [
            {"index": 0, "name": "Hoechst", "dye": "Hoechst"},
            {"index": 1, "name": "PI", "dye": "PI"},
        ]
        result = classify_channels(metadata, scenes)
        self.assertEqual([c.role for c in result], ["total", "dead"])


class CalibratedCountingTests(unittest.TestCase):
    def test_total_nuclei_have_one_label_per_peak(self):
        yy, xx = np.mgrid[:100, :100]
        image = np.zeros((100, 100), dtype=np.float32)
        for y, x in [(20, 20), (50, 55), (80, 25)]:
            image += 180 * np.exp(-((yy-y)**2 + (xx-x)**2) / (2 * 5**2))
        result = count_cells(np.clip(image, 0, 255).astype(np.uint8),
                             "fluorescence", role="total")
        self.assertEqual(result["total"], 3)
        self.assertEqual(result["labels"].max(), 3)

    def test_dead_fragments_are_merged_before_nucleus_matching(self):
        yy, xx = np.mgrid[:100, :100]
        image = np.zeros((100, 100), dtype=np.float32)
        # Two close fragments belong to one dead nucleus; one distant peak is
        # a second dead nucleus.
        for y, x in [(25, 25), (25, 34), (75, 75)]:
            image += 220 * np.exp(-((yy-y)**2 + (xx-x)**2) / (2 * 3**2))
        result = count_cells(np.clip(image, 0, 255).astype(np.uint8),
                             "fluorescence", role="dead")
        self.assertEqual(result["total"], 2)

    def test_peak_matching_uses_calibrated_distance_and_deduplicates(self):
        total = {"labels": np.zeros((50, 50), dtype=np.int32), "props": [
            {"label": 1, "peak": (10, 10)},
            {"label": 2, "peak": (35, 35)},
        ]}
        dead = {"labels": np.zeros((50, 50), dtype=np.int32), "props": [
            {"label": 1, "peak": (10, 11)},
            {"label": 2, "peak": (11, 10)},
            {"label": 3, "peak": (20, 35)},
        ]}
        self.assertEqual(match_dead_cells(total, dead), {1})


class DeadCellMatchingTests(unittest.TestCase):
    def test_only_valid_meaningful_overlaps_are_counted_once(self):
        total_labels = np.zeros((20, 20), dtype=np.int32)
        total_labels[2:8, 2:8] = 1
        total_labels[10:17, 10:17] = 2
        dead_labels = np.zeros_like(total_labels)
        dead_labels[3:7, 3:7] = 1       # good match to cell 1
        dead_labels[11:14, 11:14] = 2   # first match to cell 2
        dead_labels[14:16, 14:16] = 3   # duplicate match to cell 2
        dead_labels[0:2, 15:18] = 4     # background noise

        total = {"labels": total_labels,
                 "props": [{"label": 1}, {"label": 2}]}
        dead = {"labels": dead_labels,
                "props": [{"label": i} for i in range(1, 5)]}
        self.assertEqual(match_dead_cells(total, dead), {1, 2})

    def test_small_channel_offset_is_tolerated_but_distant_spot_is_not(self):
        total_labels = np.zeros((40, 40), dtype=np.int32)
        total_labels[10:20, 10:20] = 1
        dead_labels = np.zeros_like(total_labels)
        dead_labels[12:16, 22:25] = 1  # two-pixel gap from the cell
        dead_labels[30:34, 30:34] = 2  # unrelated background signal
        total = {"labels": total_labels, "props": [{"label": 1}]}
        dead = {"labels": dead_labels,
                "props": [{"label": 1}, {"label": 2}]}
        self.assertEqual(match_dead_cells(total, dead), {1})

    def test_rejected_watershed_labels_are_removed_and_renumbered(self):
        labels = np.array([[0, 2, 2], [5, 5, 0]], dtype=np.int32)
        props = [{"label": 5, "area": 2}]
        compact, compact_props = _compact_labels(labels, props)
        self.assertEqual(compact.max(), 1)
        self.assertEqual(int(np.count_nonzero(compact == 1)), 2)
        self.assertEqual(compact_props[0]["label"], 1)


class ExcelOutputTests(unittest.TestCase):
    def test_workbook_contains_requested_counts_only(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "result.xlsx"
            write_excel(path, [{
                "filename": "sample.czi", "scene": "Scene0",
                "total": 42, "dead": 3,
                "cell_details": [{"label": 1, "area": 10}],
            }])
            workbook = openpyxl.load_workbook(path, read_only=True)
            self.assertEqual(workbook.sheetnames, ["细胞统计"])
            sheet = workbook["细胞统计"]
            self.assertEqual([sheet.cell(1, c).value for c in range(1, 5)],
                             ["文件名", "视野", "所有细胞数", "死细胞数"])
            self.assertEqual([sheet.cell(2, c).value for c in range(1, 5)],
                             ["sample.czi", "Scene0", 42, 3])
            workbook.close()

    def test_legacy_container_values_cannot_break_excel_output(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "legacy.xlsx"
            write_excel(path, [{
                "filename": ["fluorescence", "other"],
                "scene": ("Scene0",),
                "total": [42],
                "dead": None,
            }])
            workbook = openpyxl.load_workbook(path, read_only=True)
            sheet = workbook.active
            self.assertEqual(sheet.cell(2, 1).value,
                             '["fluorescence", "other"]')
            self.assertEqual(sheet.cell(2, 3).value, 0)
            workbook.close()


if __name__ == "__main__":
    unittest.main()
